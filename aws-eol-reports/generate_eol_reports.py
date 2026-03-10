"""AWS EOL Reports Generator - Generates management and individual service Excel files."""
import os
from datetime import datetime, date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT_DIR = os.path.join(os.path.dirname(__file__), "2026-03-09")
TODAY = date(2026, 3, 9)

# ============================================================
# Style definitions
# ============================================================
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
DATA_FONT = Font(size=10)
DATA_ALIGN = Alignment(vertical="top", wrap_text=True)
CENTER_ALIGN = Alignment(horizontal="center", vertical="top", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
LABEL_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
LABEL_FONT = Font(bold=True, size=10)
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FONT = Font(color="CC0000", size=10)
BLUE_FONT = Font(color="0000CC", size=10)
ORANGE_FONT = Font(color="CC6600", size=10)
GREEN_FONT = Font(color="006600", size=10)
HYPERLINK_FONT = Font(color="0563C1", underline="single", size=10)


def style_header_row(ws, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER
    ws.auto_filter.ref = ws.dimensions


def style_data_cell(ws, row, col, center=False):
    cell = ws.cell(row=row, column=col)
    cell.font = DATA_FONT
    cell.alignment = CENTER_ALIGN if center else DATA_ALIGN
    cell.border = THIN_BORDER
    return cell


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_kv_sheet(ws, title, data, col_a_width=30, col_b_width=60):
    """Write key-value pairs sheet."""
    ws.title = title
    ws.column_dimensions["A"].width = col_a_width
    ws.column_dimensions["B"].width = col_b_width
    for r, (key, val) in enumerate(data, 1):
        ca = ws.cell(row=r, column=1, value=key)
        ca.font = LABEL_FONT
        ca.fill = LABEL_FILL
        ca.alignment = DATA_ALIGN
        ca.border = THIN_BORDER
        cb = ws.cell(row=r, column=2, value=val)
        cb.font = DATA_FONT
        cb.alignment = DATA_ALIGN
        cb.border = THIN_BORDER


def write_table_sheet(ws, title, headers, widths, rows, center_cols=None):
    """Write table with headers and rows."""
    ws.title = title
    center_cols = center_cols or []
    set_col_widths(ws, widths)
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header_row(ws, len(headers))
    for r, row_data in enumerate(rows, 2):
        for c, val in enumerate(row_data, 1):
            cell = style_data_cell(ws, r, c, center=(c in center_cols))
            cell.value = val


def write_references_sheet(ws, title, refs):
    """Write references sheet with hyperlinks."""
    ws.title = title
    headers = ["No.", "タイトル", "URL", "種別", "概要"]
    widths = [6, 40, 65, 20, 50]
    set_col_widths(ws, widths)
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header_row(ws, len(headers))
    for r, ref in enumerate(refs, 2):
        style_data_cell(ws, r, 1, center=True).value = r - 1
        style_data_cell(ws, r, 2).value = ref["title"]
        url_cell = style_data_cell(ws, r, 3)
        url_cell.value = ref["url"]
        url_cell.hyperlink = ref["url"]
        url_cell.font = HYPERLINK_FONT
        style_data_cell(ws, r, 4, center=True).value = ref["type"]
        style_data_cell(ws, r, 5).value = ref.get("desc", "")


def write_eol_steps_sheet(ws, title, steps):
    """Write EOL steps sheet."""
    ws.title = title
    headers = ["ステップNo.", "フェーズ", "作業内容", "詳細説明", "前提条件", "注意事項", "推定工数"]
    widths = [10, 15, 35, 55, 30, 30, 12]
    set_col_widths(ws, widths)
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header_row(ws, len(headers))
    for r, s in enumerate(steps, 2):
        for c, val in enumerate(s, 1):
            cell = style_data_cell(ws, r, c, center=(c in [1, 2, 7]))
            cell.value = val


def write_eol_dates_sheet(ws, title, dates_data):
    """Write EOL dates sheet."""
    ws.title = title
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 45
    headers = ["項目", "日付", "備考"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header_row(ws, 3)
    for r, (label, dt, note) in enumerate(dates_data, 2):
        ca = style_data_cell(ws, r, 1)
        ca.value = label
        ca.font = LABEL_FONT
        ca.fill = LABEL_FILL
        cb = style_data_cell(ws, r, 2, center=True)
        cb.value = dt
        if dt and dt != "未定" and dt != "N/A":
            try:
                d = datetime.strptime(dt, "%Y-%m-%d").date()
                if d < TODAY:
                    cb.font = RED_FONT
            except ValueError:
                pass
        style_data_cell(ws, r, 3).value = note


CODE_FONT = Font(name="Consolas", size=9)
CODE_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
CODE_ALIGN = Alignment(vertical="top", wrap_text=True)
SECTION_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
SECTION_FONT = Font(bold=True, color="FFFFFF", size=11)


def write_cloudshell_sheet(ws, title, commands):
    """Write CloudShell commands sheet.
    commands: list of dict with keys: category, purpose, command, notes
    """
    ws.title = title
    headers = ["No.", "カテゴリ", "用途", "CloudShellコマンド", "補足・注意事項"]
    widths = [6, 18, 30, 80, 40]
    set_col_widths(ws, widths)
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header_row(ws, len(headers))
    for r, cmd in enumerate(commands, 2):
        style_data_cell(ws, r, 1, center=True).value = r - 1
        style_data_cell(ws, r, 2, center=True).value = cmd["category"]
        style_data_cell(ws, r, 3).value = cmd["purpose"]
        # Command cell with monospace font
        cmd_cell = style_data_cell(ws, r, 4)
        cmd_cell.value = cmd["command"]
        cmd_cell.font = CODE_FONT
        cmd_cell.fill = CODE_FILL
        cmd_cell.alignment = CODE_ALIGN
        style_data_cell(ws, r, 5).value = cmd.get("notes", "")
    ws.row_dimensions[1].height = 20
    for r in range(2, len(commands) + 2):
        ws.row_dimensions[r].height = max(40, 15 * cmd["command"].count("\n") + 20)


# ============================================================
# CloudShell command definitions per service category
# ============================================================
CLOUDSHELL_COMMANDS = {
    "rds": [
        {
            "category": "RDS",
            "purpose": "RDSインスタンス一覧（エンジン・バージョン付き）",
            "command": 'aws rds describe-db-instances \\\n  --query "DBInstances[].{ID:DBInstanceIdentifier,Engine:Engine,Version:EngineVersion,Status:DBInstanceStatus,Class:DBInstanceClass}" \\\n  --output table',
            "notes": "全リージョンで実行する場合は --region を指定",
        },
        {
            "category": "RDS",
            "purpose": "特定エンジンバージョンのインスタンスをフィルタ",
            "command": "aws rds describe-db-instances \\\n  --query \"DBInstances[?Engine=='mysql'&&starts_with(EngineVersion,'5.7')].{ID:DBInstanceIdentifier,Version:EngineVersion,Class:DBInstanceClass}\" \\\n  --output table",
            "notes": "Engine名とバージョンを変更して各対象を検索\n例: 'postgres' + '13', 'mysql' + '8.0'",
        },
        {
            "category": "RDS",
            "purpose": "Extended Supportの料金確認",
            "command": 'aws rds describe-db-instances \\\n  --query "DBInstances[].{ID:DBInstanceIdentifier,Engine:Engine,Version:EngineVersion,MultiAZ:MultiAZ,VCPUs:ProcessorFeatures}" \\\n  --output table',
            "notes": "vCPU数×Extended Support単価で追加コストを算出",
        },
    ],
    "aurora": [
        {
            "category": "Aurora",
            "purpose": "Auroraクラスター一覧（エンジン・バージョン付き）",
            "command": 'aws rds describe-db-clusters \\\n  --query "DBClusters[].{ID:DBClusterIdentifier,Engine:Engine,Version:EngineVersion,Status:Status,Members:DBClusterMembers[].DBInstanceIdentifier}" \\\n  --output table',
            "notes": "Writer/Readerインスタンスも確認可能",
        },
        {
            "category": "Aurora",
            "purpose": "特定エンジンバージョンのクラスターをフィルタ",
            "command": "aws rds describe-db-clusters \\\n  --query \"DBClusters[?Engine=='aurora-postgresql'&&starts_with(EngineVersion,'13.')].{ID:DBClusterIdentifier,Version:EngineVersion}\" \\\n  --output table",
            "notes": "Engine名: 'aurora-mysql', 'aurora-postgresql'",
        },
    ],
    "elasticache": [
        {
            "category": "ElastiCache",
            "purpose": "ElastiCacheクラスター一覧",
            "command": 'aws elasticache describe-cache-clusters \\\n  --query "CacheClusters[].{ID:CacheClusterId,Engine:Engine,Version:EngineVersion,NodeType:CacheNodeType,Status:CacheClusterStatus}" \\\n  --output table',
            "notes": "Engine: redis / memcached / valkey",
        },
        {
            "category": "ElastiCache",
            "purpose": "レプリケーショングループ一覧",
            "command": 'aws elasticache describe-replication-groups \\\n  --query "ReplicationGroups[].{ID:ReplicationGroupId,Status:Status,Description:Description,NodeGroups:NodeGroups[].{Slots:Slots,Members:NodeGroupMembers[].CacheClusterId}}" \\\n  --output table',
            "notes": "クラスターモード有効な場合のシャード情報確認",
        },
    ],
    "opensearch": [
        {
            "category": "OpenSearch",
            "purpose": "OpenSearchドメイン一覧",
            "command": "aws opensearch list-domain-names --output table",
            "notes": "",
        },
        {
            "category": "OpenSearch",
            "purpose": "全ドメインの詳細（バージョン含む）",
            "command": 'for domain in $(aws opensearch list-domain-names --query "DomainNames[].DomainName" --output text); do\n  aws opensearch describe-domain --domain-name $domain \\\n    --query "DomainStatus.{Name:DomainName,Version:EngineVersion,InstanceType:ClusterConfig.InstanceType,Count:ClusterConfig.InstanceCount}" \\\n    --output table\ndone',
            "notes": "ドメインが多い場合は時間がかかる",
        },
    ],
    "lambda": [
        {
            "category": "Lambda",
            "purpose": "全Lambda関数のランタイム一覧",
            "command": 'aws lambda list-functions \\\n  --query "Functions[].{Name:FunctionName,Runtime:Runtime,LastModified:LastModified}" \\\n  --output table',
            "notes": "関数数が多い場合は --max-items で制限",
        },
        {
            "category": "Lambda",
            "purpose": "特定ランタイムの関数のみフィルタ",
            "command": "aws lambda list-functions \\\n  --query \"Functions[?Runtime=='python3.9'].{Name:FunctionName,Runtime:Runtime,LastModified:LastModified}\" \\\n  --output table",
            "notes": "Runtime値を変更して各対象を検索\n例: 'python3.10', 'nodejs20.x', 'ruby3.2', 'dotnet8', 'provided.al2'",
        },
        {
            "category": "Lambda",
            "purpose": "ランタイム別の関数数カウント",
            "command": 'aws lambda list-functions \\\n  --query "Functions[].Runtime" \\\n  --output text | tr \'\\t\' \'\\n\' | sort | uniq -c | sort -rn',
            "notes": "ランタイム別の使用状況を俯瞰する際に有用",
        },
    ],
    "ec2": [
        {
            "category": "EC2/AL2",
            "purpose": "稼働中EC2インスタンスのAMI情報",
            "command": 'aws ec2 describe-instances \\\n  --filters "Name=instance-state-name,Values=running" \\\n  --query "Reservations[].Instances[].{ID:InstanceId,Type:InstanceType,AMI:ImageId,Name:Tags[?Key==\'Name\']|[0].Value}" \\\n  --output table',
            "notes": "AMI IDからAL2かどうかを判定する",
        },
        {
            "category": "EC2/AL2",
            "purpose": "使用中AMIの詳細確認（AL2判定）",
            "command": 'for ami in $(aws ec2 describe-instances \\\n  --filters "Name=instance-state-name,Values=running" \\\n  --query "Reservations[].Instances[].ImageId" \\\n  --output text | sort -u); do\n  aws ec2 describe-images --image-ids $ami \\\n    --query "Images[].{AMI:ImageId,Name:Name,Description:Description}" \\\n    --output table 2>/dev/null\ndone',
            "notes": "Name/Descriptionに 'amzn2' が含まれればAL2",
        },
        {
            "category": "EC2/AL2",
            "purpose": "AL2 AMIのみフィルタ（名前ベース）",
            "command": 'aws ec2 describe-instances \\\n  --filters "Name=instance-state-name,Values=running" \\\n  --query "Reservations[].Instances[].{ID:InstanceId,AMI:ImageId,Name:Tags[?Key==\'Name\']|[0].Value}" \\\n  --output json | python3 -c "\nimport json,sys,subprocess\ninstances=json.load(sys.stdin)\nfor flat in [i for sub in instances for i in (sub if isinstance(sub,list) else [sub])]:\n  ami=flat.get(\'AMI\',\'\')\n  r=subprocess.run([\'aws\',\'ec2\',\'describe-images\',\'--image-ids\',ami,\'--query\',\'Images[0].Name\',\'--output\',\'text\'],capture_output=True,text=True)\n  if \'amzn2\' in r.stdout.lower():\n    print(f\\\"AL2: {flat.get(\'ID\')} ({flat.get(\'Name\',\'-\')}) AMI={ami}\\\")"',
            "notes": "AL2インスタンスのみ抽出するスクリプト\nCloudShellにはPython3が標準装備",
        },
    ],
    "windows_server": [
        {
            "category": "Windows Server",
            "purpose": "Windowsプラットフォームの稼働中EC2一覧",
            "command": 'aws ec2 describe-instances \\\n  --filters "Name=instance-state-name,Values=running" "Name=platform-details,Values=Windows*" \\\n  --query "Reservations[].Instances[].{ID:InstanceId,Type:InstanceType,AMI:ImageId,Platform:PlatformDetails,Name:Tags[?Key==\'Name\']|[0].Value}" \\\n  --output table',
            "notes": "platform-detailsフィルタでWindows Serverのみ抽出",
        },
        {
            "category": "Windows Server",
            "purpose": "Windows AMIのOS名・バージョン確認",
            "command": 'for ami in $(aws ec2 describe-instances \\\n  --filters "Name=instance-state-name,Values=running" "Name=platform-details,Values=Windows*" \\\n  --query "Reservations[].Instances[].ImageId" \\\n  --output text | sort -u); do\n  aws ec2 describe-images --image-ids $ami \\\n    --query "Images[].{AMI:ImageId,Name:Name}" \\\n    --output table 2>/dev/null\ndone',
            "notes": "AMI名に 'Windows_Server-2012' 'Windows_Server-2016' が含まれるものがEOL対象",
        },
        {
            "category": "Windows Server",
            "purpose": "SSM Inventoryで詳細OS情報を取得",
            "command": 'aws ssm describe-instance-information \\\n  --filters "Key=PlatformTypes,Values=Windows" \\\n  --query "InstanceInformationList[].{ID:InstanceId,Name:ComputerName,OS:PlatformName,Version:PlatformVersion,AgentVersion:AgentVersion}" \\\n  --output table',
            "notes": "SSM Agentがインストール済みのインスタンスのみ対象\nPlatformVersionでOSバージョンを正確に判別可能",
        },
        {
            "category": "Windows Server",
            "purpose": "Trusted AdvisorでWindows EOS検出",
            "command": 'aws support describe-trusted-advisor-checks \\\n  --language en \\\n  --query "checks[?name==\'Amazon EC2 instances running Microsoft Windows Server end of support\'].{ID:id,Name:name}" \\\n  --output table',
            "notes": "Business/Enterprise Supportプラン必要\nTrusted AdvisorがEOS対象インスタンスを自動検出",
        },
    ],
    "multi_account": [
        {
            "category": "マルチアカウント",
            "purpose": "Organizations全アカウント一覧",
            "command": 'aws organizations list-accounts \\\n  --query "Accounts[?Status==\'ACTIVE\'].{ID:Id,Name:Name,Email:Email}" \\\n  --output table',
            "notes": "管理アカウントでのみ実行可能",
        },
        {
            "category": "マルチアカウント",
            "purpose": "他アカウントへのスイッチロール",
            "command": 'CREDS=$(aws sts assume-role \\\n  --role-arn "arn:aws:iam::TARGET_ACCOUNT_ID:role/ReadOnlyRole" \\\n  --role-session-name "eol-audit" \\\n  --query "Credentials" --output json)\n\nexport AWS_ACCESS_KEY_ID=$(echo $CREDS | python3 -c "import sys,json;print(json.load(sys.stdin)[\'AccessKeyId\'])")\nexport AWS_SECRET_ACCESS_KEY=$(echo $CREDS | python3 -c "import sys,json;print(json.load(sys.stdin)[\'SecretAccessKey\'])")\nexport AWS_SESSION_TOKEN=$(echo $CREDS | python3 -c "import sys,json;print(json.load(sys.stdin)[\'SessionToken\'])")',
            "notes": "TARGET_ACCOUNT_IDとロール名を環境に合わせて変更\n実行後は上記の各コマンドがターゲットアカウントに対して実行される",
        },
        {
            "category": "マルチアカウント",
            "purpose": "全リージョンでコマンドを一括実行",
            "command": 'for region in $(aws ec2 describe-regions --query "Regions[].RegionName" --output text); do\n  echo "=== $region ==="\n  aws rds describe-db-instances --region $region \\\n    --query "DBInstances[].{ID:DBInstanceIdentifier,Engine:Engine,Version:EngineVersion}" \\\n    --output table 2>/dev/null\ndone',
            "notes": "rds以外のコマンドに差し替えて各サービスで使用可能\n全リージョンのスキャンには数分かかる",
        },
    ],
}

# Mapping: service filename -> which CloudShell command categories to include
SERVICE_TO_CLOUDSHELL = {
    "opensearch-legacy": ["opensearch", "multi_account"],
    "lambda-python-3.9": ["lambda", "multi_account"],
    "elasticache-redis-4-5": ["elasticache", "multi_account"],
    "rds-postgresql-13": ["rds", "multi_account"],
    "aurora-postgresql-13": ["aurora", "multi_account"],
    "lambda-ruby-3.2": ["lambda", "multi_account"],
    "lambda-nodejs-20": ["lambda", "multi_account"],
    "amazon-linux-2": ["ec2", "multi_account"],
    "lambda-custom-runtime-al2": ["lambda", "multi_account"],
    "rds-mysql-8.0": ["rds", "multi_account"],
    "lambda-python-3.10": ["lambda", "multi_account"],
    "lambda-dotnet-8": ["lambda", "multi_account"],
    "aws-sdk-dotnet-v3": ["multi_account"],
    "elasticache-redis-6": ["elasticache", "multi_account"],
    "rds-aurora-postgresql-14": ["rds", "aurora", "multi_account"],
    "rds-mysql-5.7-extended": ["rds", "multi_account"],
    "aurora-mysql-2.x-extended": ["aurora", "multi_account"],
    "lambda-2027-runtimes": ["lambda", "multi_account"],
    "windows-server-2012": ["windows_server", "multi_account"],
    "windows-server-2016": ["windows_server", "multi_account"],
}


def get_cloudshell_commands_for_service(filename):
    """Get relevant CloudShell commands for a specific service."""
    categories = SERVICE_TO_CLOUDSHELL.get(filename, ["multi_account"])
    commands = []
    for cat in categories:
        commands.extend(CLOUDSHELL_COMMANDS.get(cat, []))
    return commands


def days_until(date_str):
    try:
        d = datetime.strptime(date_str, "%Y-%m-%d").date()
        return (d - TODAY).days
    except (ValueError, TypeError):
        return None


def priority_from_days(days):
    if days is None:
        return "中"
    if days < 90:
        return "高"
    elif days < 180:
        return "中"
    return "低"


def save_wb(wb, filename):
    path = os.path.join(OUTPUT_DIR, filename)
    wb.save(path)
    print(f"  Created: {filename}")


# ============================================================
# Service data
# ============================================================
SERVICES = [
    {
        "no": 1, "name": "OpenSearch (Legacy Versions)", "target": "ES~6.7, 7.1~7.8, OS 1.0~2.9",
        "current": "Elasticsearch 6.x-7.x / OpenSearch 1.x-2.x", "eol_date": "2025-11-07",
        "ext_support_end": "2026-11-07", "migration_to": "OpenSearch 2.11以降",
        "priority": "高", "status": "未着手", "notes": "標準サポート終了済み、Extended Support中",
        "filename": "opensearch-legacy",
        "summary": [
            ("サービス名", "Amazon OpenSearch Service"),
            ("EOL対象バージョン", "Elasticsearch ~6.7, 7.1~7.8 / OpenSearch 1.0~1.2, 2.3~2.9"),
            ("AWS公式アナウンス日", "2024-11"),
            ("標準サポート終了日", "2025-11-07"),
            ("延長サポート終了日", "2026-11-07"),
            ("延長サポート費用", "$0.0065/NIH（US East）"),
            ("影響範囲", "セキュリティパッチ停止、新機能なし"),
            ("セキュリティへの影響", "標準サポート終了後はExtended Supportで重要パッチのみ提供"),
            ("代替サービス/移行先", "OpenSearch 2.11以降"),
            ("移行の複雑度", "中（ブルー/グリーンデプロイで対応可能）"),
            ("推定ダウンタイム", "ブルー/グリーンデプロイにより最小化可能"),
            ("追加コスト", "Extended Support: $0.0065/NIH（US East）"),
        ],
        "steps": [
            (1, "準備", "現行バージョンの確認", "全ドメインのエンジンバージョンを確認し一覧化する", "AWSコンソールアクセス", "本番・検証全環境を対象とすること", "0.5日"),
            (2, "準備", "互換性の確認", "現行バージョンと移行先バージョン間のAPI互換性を確認", "移行先バージョンの決定", "Breaking Changesの有無を確認", "1日"),
            (3, "準備", "インデックスマッピングの確認", "非推奨のマッピングタイプやフィールドがないか確認", "なし", "Elasticsearch 6.x以前のマッピングは要変更", "1日"),
            (4, "検証", "検証環境でのアップグレードテスト", "スナップショットから検証環境を作成しアップグレード実施", "検証環境の準備", "クエリの動作確認を必ず実施", "2-3日"),
            (5, "実施", "ブルー/グリーンデプロイでアップグレード", "本番ドメインのブルー/グリーンデプロイを実施", "検証環境テスト完了", "ロールバック手順を事前に確認", "1-2日"),
            (6, "確認", "動作確認", "アップグレード後のクエリ動作・パフォーマンスを確認", "アップグレード完了", "インデックスの再作成が必要な場合あり", "1日"),
        ],
        "dates": [
            ("AWS標準サポート終了日", "2025-11-07", "既に終了済み"),
            ("延長サポート終了日", "2026-11-07", "追加費用発生中"),
            ("推奨対応開始日", "2025-08-07", "EOL 3ヶ月前"),
            ("推奨対応完了日", "2026-08-07", "延長サポート終了3ヶ月前"),
        ],
        "refs": [
            {"title": "OpenSearch Standard/Extended Support発表", "url": "https://aws.amazon.com/blogs/big-data/amazon-opensearch-service-announces-standard-and-extended-support-dates-for-elasticsearch-and-opensearch-versions/", "type": "AWSブログ", "desc": "標準・延長サポート日程の公式発表"},
            {"title": "OpenSearch Extended Support発表", "url": "https://aws.amazon.com/about-aws/whats-new/2024/11/amazon-opensearch-service-support-engine-versions/", "type": "AWS What's New", "desc": "Extended Support機能のリリース告知"},
            {"title": "OpenSearch バージョンEOLスケジュール", "url": "https://endoflife.date/opensearch", "type": "サードパーティ", "desc": "バージョン別EOL一覧"},
        ],
    },
    {
        "no": 2, "name": "Lambda Python 3.9", "target": "ランタイム",
        "current": "Python 3.9", "eol_date": "2025-12-15",
        "ext_support_end": "2026-02-15", "migration_to": "Python 3.12 / 3.13",
        "priority": "高", "status": "未着手", "notes": "Phase 1開始済み（新規作成不可）",
        "filename": "lambda-python-3.9",
        "summary": [
            ("サービス名", "AWS Lambda - Python 3.9 ランタイム"),
            ("EOL対象", "Python 3.9 ランタイム"),
            ("Phase 1（新規作成不可）", "2025-12-15"),
            ("Phase 2a（更新不可）", "2026-02-15"),
            ("延長サポート", "なし（Lambdaランタイムに延長サポートはない）"),
            ("影響範囲", "Phase 1: 新規Lambda関数の作成不可\nPhase 2: 既存関数の更新不可（実行は継続可能）"),
            ("セキュリティへの影響", "セキュリティパッチの提供停止"),
            ("代替サービス/移行先", "Python 3.12 または 3.13"),
            ("移行の複雑度", "低〜中（依存ライブラリの互換性確認が必要）"),
            ("推定ダウンタイム", "なし（新バージョンのデプロイで切り替え）"),
            ("追加コスト", "なし"),
        ],
        "steps": [
            (1, "準備", "対象Lambda関数の洗い出し", "Python 3.9を使用している全Lambda関数を一覧化", "AWSコンソールアクセス", "全リージョンを確認すること", "0.5日"),
            (2, "準備", "依存ライブラリの互換性確認", "requirements.txtの各ライブラリがPython 3.12/3.13で動作するか確認", "対象関数一覧", "特にnumpy, pandas等のネイティブライブラリに注意", "1日"),
            (3, "検証", "検証環境でのテスト", "Python 3.12/3.13ランタイムで関数をデプロイしテスト実行", "互換性確認完了", "レイヤーの互換性も確認", "1-2日"),
            (4, "実施", "ランタイムバージョンの更新", "Lambda関数のランタイムをPython 3.12/3.13に変更してデプロイ", "テスト完了", "CloudFormation/SAM/CDKのテンプレートも更新", "0.5日/関数"),
            (5, "確認", "動作確認", "本番環境での動作確認・ログ監視", "デプロイ完了", "エラー発生時のロールバック手順を準備", "0.5日"),
        ],
        "dates": [
            ("Phase 1開始（新規作成不可）", "2025-12-15", "既に開始済み"),
            ("Phase 2a（更新不可）", "2026-02-15", "既存関数の更新不可"),
            ("推奨対応完了日", "2025-12-01", "Phase 1前に対応推奨"),
        ],
        "refs": [
            {"title": "Lambda ランタイム一覧", "url": "https://docs.aws.amazon.com/lambda/latest/dg/lambda-runtimes.html", "type": "AWS公式ドキュメント", "desc": "全ランタイムのサポートスケジュール"},
            {"title": "AWS Lambda EOL一覧", "url": "https://endoflife.date/aws-lambda", "type": "サードパーティ", "desc": "全ランタイムのEOLスケジュール"},
        ],
    },
    {
        "no": 3, "name": "ElastiCache Redis 4.x/5.x", "target": "エンジンバージョン",
        "current": "Redis 4.x / 5.x", "eol_date": "2026-01-31",
        "ext_support_end": "2029-01-31", "migration_to": "ElastiCache for Valkey 8.0 / Redis 7.x",
        "priority": "高", "status": "未着手", "notes": "Extended Support自動登録、+80%プレミアム",
        "filename": "elasticache-redis-4-5",
        "summary": [
            ("サービス名", "Amazon ElastiCache for Redis OSS 4.x / 5.x"),
            ("EOL対象バージョン", "Redis OSS 4.x および 5.x"),
            ("標準サポート終了日", "2026-01-31"),
            ("延長サポート", "あり（2026-02-01〜2029-01-31、最大3年）"),
            ("延長サポート費用", "Year 1-2: On-Demand価格の+80%\nYear 3: On-Demand価格の+160%"),
            ("影響範囲", "セキュリティパッチ・バグ修正は延長サポートで提供（重要なもののみ）"),
            ("セキュリティへの影響", "標準サポート終了後はExtended Supportの重要パッチのみ"),
            ("代替サービス/移行先", "ElastiCache for Valkey 8.0（推奨）またはRedis 7.x"),
            ("移行の複雑度", "中（Valkeyはdrop-in replacement）"),
            ("推定ダウンタイム", "インプレースアップグレードで数分程度"),
            ("追加コスト", "Valkey移行時: コスト削減（最大33%減）\nExtended Support: +80%〜160%増"),
        ],
        "steps": [
            (1, "準備", "現行クラスターの棚卸し", "Redis 4.x/5.xを使用中の全クラスターを一覧化", "AWSコンソールアクセス", "全リージョン確認", "0.5日"),
            (2, "準備", "Valkey互換性の確認", "使用中のRedisコマンド・機能がValkeyでサポートされるか確認", "クラスター一覧", "LuaスクリプトやModulesの互換性に注意", "1日"),
            (3, "準備", "レプリカノード追加", "各シャードにレプリカノードを1つ追加（ダウンタイム最小化のため）", "なし", "追加費用が発生", "0.5日"),
            (4, "検証", "検証環境でのアップグレードテスト", "検証クラスターでRedis→Valkeyのインプレースアップグレードを実施", "検証環境準備", "データ整合性の確認", "1日"),
            (5, "実施", "本番クラスターのアップグレード", "インプレースエンジンアップグレードを実施（Redis→Valkey）", "テスト完了", "Redis <5.0.6からの場合フェイルオーバー30-60秒", "1日"),
            (6, "確認", "動作確認・パフォーマンス検証", "アプリケーションの接続・パフォーマンスを確認", "アップグレード完了", "レイテンシ・スループットの変化を監視", "0.5日"),
        ],
        "dates": [
            ("標準サポート終了日", "2026-01-31", "この日以降Extended Support自動登録"),
            ("延長サポート Year 1-2", "2026-02-01", "+80%プレミアム開始"),
            ("延長サポート Year 3", "2028-02-01", "+160%プレミアム開始"),
            ("延長サポート終了日", "2029-01-31", "強制アップグレード"),
            ("推奨対応開始日", "2025-10-31", "EOL 3ヶ月前"),
            ("推奨対応完了日", "2026-01-01", "EOL 1ヶ月前"),
        ],
        "refs": [
            {"title": "ElastiCache Redis 4/5 Extended Support発表", "url": "https://aws.amazon.com/blogs/database/introducing-extended-support-for-amazon-elasticache-version-4-and-version-5-for-redis-oss/", "type": "AWSブログ", "desc": "Extended Supportの詳細と価格"},
            {"title": "ElastiCache Extended Supportドキュメント", "url": "https://docs.aws.amazon.com/AmazonElastiCache/latest/dg/extended-support.html", "type": "AWS公式ドキュメント", "desc": "Extended Supportの概要"},
            {"title": "ElastiCache Extended Support対象バージョン", "url": "https://docs.aws.amazon.com/AmazonElastiCache/latest/dg/extended-support-versions.html", "type": "AWS公式ドキュメント", "desc": "対象バージョン一覧"},
            {"title": "ElastiCache Extended Support料金", "url": "https://docs.aws.amazon.com/AmazonElastiCache/latest/dg/extended-support-charges.html", "type": "AWS公式ドキュメント", "desc": "料金体系"},
            {"title": "エンジンバージョンアップグレード", "url": "https://docs.aws.amazon.com/AmazonElastiCache/latest/dg/VersionManagement.HowTo.html", "type": "AWS公式ドキュメント", "desc": "クロスエンジンアップグレード手順"},
            {"title": "Redis→Valkeyアップグレードガイド", "url": "https://repost.aws/knowledge-center/elasticache-upgrade-redis-to-valkey", "type": "AWS re:Post", "desc": "Valkeyへの移行手順"},
        ],
    },
    {
        "no": 4, "name": "RDS PostgreSQL 13.x", "target": "メジャーバージョン",
        "current": "PostgreSQL 13.x", "eol_date": "2026-02-28",
        "ext_support_end": "2029-02-28", "migration_to": "PostgreSQL 16以上",
        "priority": "高", "status": "未着手", "notes": "Extended Support: $0.100/vCPU-hr (Y1-2)",
        "filename": "rds-postgresql-13",
        "summary": [
            ("サービス名", "Amazon RDS for PostgreSQL 13.x"),
            ("EOL対象バージョン", "PostgreSQL 13.x"),
            ("コミュニティEOL", "2025-11"),
            ("AWS標準サポート終了日", "2026-02-28"),
            ("延長サポート", "あり（2026-03-01〜2029-02-28、最大3年）"),
            ("延長サポート費用", "Year 1-2: $0.100/vCPU-hr\nYear 3: $0.200/vCPU-hr"),
            ("影響範囲", "標準サポート終了後は新しいマイナーバージョンリリースなし"),
            ("セキュリティへの影響", "Extended Supportで重要なCVEパッチのみ提供"),
            ("代替サービス/移行先", "PostgreSQL 16以上"),
            ("移行の複雑度", "中（Blue/Greenデプロイで最小ダウンタイム対応可能）"),
            ("推定ダウンタイム", "Blue/Greenデプロイ: 数分\nインプレース: 数十分〜数時間"),
            ("追加コスト", "Extended Support: $0.100〜0.200/vCPU-hr"),
        ],
        "steps": [
            (1, "準備", "対象インスタンスの棚卸し", "PG 13.xを使用中の全RDSインスタンスを一覧化", "AWSコンソールアクセス", "Multi-AZ・Read Replicaも含めて確認", "0.5日"),
            (2, "準備", "互換性の事前確認", "PG 13→16間の非互換項目（拡張機能、SQL構文等）を調査", "対象一覧", "pg_upgradeの事前チェック", "1日"),
            (3, "準備", "パラメータグループの準備", "PG 16用のカスタムパラメータグループを作成", "なし", "PG 13固有のパラメータを確認", "0.5日"),
            (4, "準備", "論理レプリケーション有効化", "rds.logical_replication=1を設定し再起動", "Blue/Green使用時", "再起動が必要", "0.5日"),
            (5, "検証", "Blue/Greenデプロイテスト", "検証環境でBlue/Greenデプロイを使ってPG 16にアップグレード", "検証環境準備", "全テーブルにプライマリキーが必要", "2日"),
            (6, "実施", "本番Blue/Greenデプロイ", "本番環境でBlue/Greenデプロイを実施しスイッチオーバー", "テスト完了", "スイッチオーバーのタイミングを計画", "1日"),
            (7, "確認", "動作確認・パフォーマンス検証", "クエリ実行計画・パフォーマンスの確認", "スイッチオーバー完了", "Greenの旧環境は保持して問題時にロールバック", "1日"),
        ],
        "dates": [
            ("コミュニティEOL", "2025-11-01", "PostgreSQLコミュニティのサポート終了"),
            ("AWS標準サポート終了日", "2026-02-28", "Extended Support自動登録"),
            ("延長サポート Year 1-2", "2026-03-01", "$0.100/vCPU-hr"),
            ("延長サポート Year 3", "2028-03-01", "$0.200/vCPU-hr"),
            ("延長サポート終了日", "2029-02-28", "強制メジャーバージョンアップグレード"),
            ("推奨対応開始日", "2025-11-28", "EOL 3ヶ月前"),
            ("推奨対応完了日", "2026-01-28", "EOL 1ヶ月前"),
        ],
        "refs": [
            {"title": "RDS PostgreSQL 13 EOL発表", "url": "https://repost.aws/articles/ARRvHxJ_9sTDCGloBavca3kg/announcement-amazon-rds-postgresql-13-x-end-of-standard-support-is-february-28-2026", "type": "AWS re:Post", "desc": "公式EOLアナウンス"},
            {"title": "RDS PostgreSQLリリースカレンダー", "url": "https://docs.aws.amazon.com/AmazonRDS/latest/PostgreSQLReleaseNotes/postgresql-release-calendar.html", "type": "AWS公式ドキュメント", "desc": "全バージョンのサポートスケジュール"},
            {"title": "RDS Extended Support", "url": "https://docs.aws.amazon.com/AmazonRDS/latest/UserGuide/extended-support.html", "type": "AWS公式ドキュメント", "desc": "Extended Supportの概要"},
            {"title": "PostgreSQL 13アップグレード戦略", "url": "https://aws.amazon.com/blogs/database/strategies-for-upgrading-amazon-aurora-postgresql-and-amazon-rds-for-postgresql-from-version-13/", "type": "AWSブログ", "desc": "アップグレード方法の比較"},
            {"title": "Blue/Greenデプロイ", "url": "https://docs.aws.amazon.com/AmazonRDS/latest/UserGuide/blue-green-deployments.html", "type": "AWS公式ドキュメント", "desc": "Blue/Greenデプロイの詳細手順"},
            {"title": "メジャーバージョンアップグレード", "url": "https://docs.aws.amazon.com/AmazonRDS/latest/UserGuide/USER_UpgradeDBInstance.PostgreSQL.html", "type": "AWS公式ドキュメント", "desc": "インプレースアップグレード手順"},
        ],
    },
    {
        "no": 5, "name": "Aurora PostgreSQL 13.x", "target": "メジャーバージョン",
        "current": "PostgreSQL 13.x", "eol_date": "2026-02-28",
        "ext_support_end": "2029-02-28", "migration_to": "Aurora PostgreSQL 16以上",
        "priority": "高", "status": "未着手", "notes": "RDS PG 13と同日程",
        "filename": "aurora-postgresql-13",
    },
    {
        "no": 6, "name": "Lambda Ruby 3.2", "target": "ランタイム",
        "current": "Ruby 3.2", "eol_date": "2026-03-31",
        "ext_support_end": "N/A", "migration_to": "Ruby 3.3 / 3.4",
        "priority": "中", "status": "未着手", "notes": "Phase 2: 2026-09-30",
        "filename": "lambda-ruby-3.2",
    },
    {
        "no": 7, "name": "Lambda Node.js 20.x", "target": "ランタイム",
        "current": "Node.js 20.x", "eol_date": "2026-04-30",
        "ext_support_end": "N/A", "migration_to": "Node.js 22",
        "priority": "中", "status": "未着手", "notes": "Phase 2a: 2026-06-01, Phase 2b: 2026-07-01",
        "filename": "lambda-nodejs-20",
    },
    {
        "no": 8, "name": "Amazon Linux 2", "target": "OS全体",
        "current": "Amazon Linux 2", "eol_date": "2026-06-30",
        "ext_support_end": "N/A", "migration_to": "Amazon Linux 2023 (AL2023)",
        "priority": "高", "status": "未着手", "notes": "EC2, ECS, EKS, Lambda等広範囲に影響",
        "filename": "amazon-linux-2",
        "summary": [
            ("サービス名", "Amazon Linux 2"),
            ("EOL対象", "Amazon Linux 2 OS全体"),
            ("EOL日", "2026-06-30"),
            ("延長サポート", "なし"),
            ("影響範囲", "EC2インスタンス、ECS/EKSノード、Lambda (provided.al2)、\nElastic Beanstalk環境等、AL2ベースの全リソース"),
            ("セキュリティへの影響", "セキュリティパッチ・カーネルアップデートの提供完全停止"),
            ("代替サービス/移行先", "Amazon Linux 2023 (AL2023)"),
            ("移行の複雑度", "高（OSレベルの変更のため広範囲なテストが必要）"),
            ("推定ダウンタイム", "新規インスタンス作成→切り替えのため、計画的な対応で最小化可能"),
            ("追加コスト", "AL2023自体の追加コストなし"),
            ("Kernel Live Patching終了", "AL2 Kernel 4.14: 2025-10-31"),
            ("Java 7サポート終了", "2026-02-01（java-1.7.0-openjdkパッケージ）"),
        ],
        "steps": [
            (1, "準備", "AL2インスタンスの棚卸し", "AL2を使用中の全EC2/ECS/EKSリソースを一覧化", "AWSコンソールアクセス", "AMI IDでフィルタリングが有効", "1日"),
            (2, "準備", "AL2023との差分確認", "パッケージマネージャ(yum→dnf)、カーネル、systemd等の差分を確認", "なし", "AL2023はインプレースアップグレード不可", "2日"),
            (3, "準備", "AL2023用AMI/コンテナイメージの準備", "AL2023ベースのAMI作成またはコンテナイメージのビルド", "差分確認完了", "Packer等でAMI作成を自動化推奨", "2日"),
            (4, "検証", "検証環境でのテスト", "AL2023ベースのインスタンスでアプリケーション動作テスト", "AMI準備完了", "全サービスの結合テストを実施", "3-5日"),
            (5, "実施", "本番環境のローリングアップデート", "Auto Scaling Groupの起動テンプレートを更新し段階的に入替", "テスト完了", "Blue/Greenデプロイパターン推奨", "2-3日"),
            (6, "確認", "動作確認・監視強化", "切り替え後のメトリクス・ログを監視", "デプロイ完了", "ロールバック用に旧AMIを保持", "1日"),
        ],
        "dates": [
            ("Kernel 4.14 Live Patching終了", "2025-10-31", "Kernel 5.10への移行推奨"),
            ("Java 7 (java-1.7.0-openjdk) 終了", "2026-02-01", "Java 8以上に移行"),
            ("Amazon Linux 2 EOL", "2026-06-30", "セキュリティパッチ完全停止"),
            ("推奨対応開始日", "2026-03-30", "EOL 3ヶ月前"),
            ("推奨対応完了日", "2026-05-30", "EOL 1ヶ月前"),
        ],
        "refs": [
            {"title": "Amazon Linux 2 FAQ", "url": "https://aws.amazon.com/amazon-linux-2/faqs/", "type": "AWS公式ドキュメント", "desc": "AL2のFAQとEOL情報"},
            {"title": "Amazon Linux EOL一覧", "url": "https://endoflife.date/amazon-linux", "type": "サードパーティ", "desc": "バージョン別EOLスケジュール"},
            {"title": "AL2→AL2023移行ガイド (Elastic Beanstalk)", "url": "https://docs.aws.amazon.com/elasticbeanstalk/latest/dg/using-features.migration-al.generic.from-al2.html", "type": "AWS公式ドキュメント", "desc": "Elastic BeanstalkでのAL2023移行"},
            {"title": "AL2→AL2023移行ガイド (ECS)", "url": "https://docs.aws.amazon.com/AmazonECS/latest/developerguide/al2-to-al2023-ami-transition.html", "type": "AWS公式ドキュメント", "desc": "ECSでのAL2023移行"},
            {"title": "Amazon Linux 2 EOL (re:Post)", "url": "https://repost.aws/questions/QU8_7ivy19Q7Wq3CKUE5b7Jw/amazon-linux-2-motd-says-eol-is-2025-06-30", "type": "AWS re:Post", "desc": "EOL日程に関するQ&A"},
        ],
    },
    {
        "no": 9, "name": "Lambda Custom Runtime (AL2)", "target": "ランタイム",
        "current": "provided.al2", "eol_date": "2026-07-31",
        "ext_support_end": "N/A", "migration_to": "provided.al2023",
        "priority": "中", "status": "未着手", "notes": "AL2 EOLに連動",
        "filename": "lambda-custom-runtime-al2",
    },
    {
        "no": 10, "name": "RDS MySQL 8.0", "target": "メジャーバージョン",
        "current": "MySQL 8.0", "eol_date": "2026-07-31",
        "ext_support_end": "2029-07-31", "migration_to": "MySQL 8.4 LTS",
        "priority": "中", "status": "未着手", "notes": "Extended Support: $0.100/vCPU-hr (Y1-2)",
        "filename": "rds-mysql-8.0",
        "summary": [
            ("サービス名", "Amazon RDS for MySQL 8.0"),
            ("EOL対象バージョン", "MySQL 8.0 メジャーバージョン"),
            ("コミュニティEOL", "2026-04-30"),
            ("AWS標準サポート終了日", "2026-07-31"),
            ("延長サポート", "あり（2026-08-01〜2029-07-31、最大3年）"),
            ("延長サポート費用", "Year 1-2: $0.100/vCPU-hr\nYear 3: $0.200/vCPU-hr"),
            ("影響範囲", "標準サポート終了後は新しいマイナーバージョンリリースなし"),
            ("セキュリティへの影響", "Extended Supportで重要なCVEパッチのみ提供"),
            ("代替サービス/移行先", "MySQL 8.4 LTS"),
            ("移行の複雑度", "中（mysql_native_password廃止に注意）"),
            ("推定ダウンタイム", "Blue/Greenデプロイ: 数分\nインプレース: テーブル再構築により数時間〜数日"),
            ("追加コスト", "Extended Support: $0.100〜0.200/vCPU-hr"),
        ],
        "steps": [
            (1, "準備", "対象インスタンスの棚卸し", "MySQL 8.0を使用中の全RDSインスタンスを一覧化", "AWSコンソールアクセス", "マイナーバージョンも確認（8.0.22未満は中間アップグレード要）", "0.5日"),
            (2, "準備", "認証プラグインの確認", "mysql_native_passwordを使用中のユーザーを確認", "なし", "8.4ではデフォルト無効のためcaching_sha2_passwordに移行必要", "1日"),
            (3, "準備", "互換性の事前チェック", "8.0→8.4間の非互換項目を確認（予約語、デフォルト値変更等）", "なし", "AWSが自動実行するprecheckの項目を事前確認", "1日"),
            (4, "検証", "Blue/Greenデプロイテスト", "検証環境でBlue/Greenデプロイを使ってMySQL 8.4にアップグレード", "検証環境準備", "テーブル再構築に要する時間を計測", "2-3日"),
            (5, "実施", "本番Blue/Greenデプロイ", "本番環境でBlue/Greenデプロイを実施しスイッチオーバー", "テスト完了", "データ量が大きい場合は長時間になる可能性", "1-2日"),
            (6, "確認", "動作確認", "アプリケーション接続・クエリ実行を確認", "スイッチオーバー完了", "認証周りの問題に注意", "1日"),
        ],
        "dates": [
            ("MySQLコミュニティEOL", "2026-04-30", "コミュニティサポート終了"),
            ("AWS標準サポート終了日", "2026-07-31", "Extended Support自動登録"),
            ("延長サポート Year 1-2", "2026-08-01", "$0.100/vCPU-hr"),
            ("延長サポート Year 3", "2028-08-01", "$0.200/vCPU-hr"),
            ("延長サポート終了日", "2029-07-31", "強制メジャーバージョンアップグレード"),
            ("推奨対応開始日", "2026-04-30", "EOL 3ヶ月前"),
            ("推奨対応完了日", "2026-06-30", "EOL 1ヶ月前"),
        ],
        "refs": [
            {"title": "RDS MySQLバージョン管理", "url": "https://docs.aws.amazon.com/AmazonRDS/latest/UserGuide/MySQL.Concepts.VersionMgmt.html", "type": "AWS公式ドキュメント", "desc": "バージョンポリシーとサポートスケジュール"},
            {"title": "RDS MySQL 8.4 LTS GA発表", "url": "https://aws.amazon.com/blogs/database/amazon-rds-for-mysql-lts-version-8-4-is-now-generally-available/", "type": "AWSブログ", "desc": "MySQL 8.4 LTSの発表とアップグレードガイド"},
            {"title": "RDS MySQLメジャーバージョンアップグレード", "url": "https://docs.aws.amazon.com/AmazonRDS/latest/UserGuide/USER_UpgradeDBInstance.MySQL.Major.html", "type": "AWS公式ドキュメント", "desc": "アップグレード手順の詳細"},
            {"title": "RDS Extended Support", "url": "https://docs.aws.amazon.com/AmazonRDS/latest/UserGuide/extended-support.html", "type": "AWS公式ドキュメント", "desc": "Extended Supportの概要"},
            {"title": "Extended Support費用見積もり", "url": "https://aws.amazon.com/blogs/aws-cloud-financial-management/estimating-the-charges-for-amazon-rds-extended-support/", "type": "AWSブログ", "desc": "Extended Support費用の算出方法"},
        ],
    },
    {
        "no": 11, "name": "Lambda Python 3.10", "target": "ランタイム",
        "current": "Python 3.10", "eol_date": "2026-10-31",
        "ext_support_end": "N/A", "migration_to": "Python 3.12 / 3.13",
        "priority": "低", "status": "未着手", "notes": "Phase 2: 2027-01-15",
        "filename": "lambda-python-3.10",
    },
    {
        "no": 12, "name": "Lambda .NET 8", "target": "ランタイム",
        "current": ".NET 8", "eol_date": "2026-11-10",
        "ext_support_end": "N/A", "migration_to": ".NET 10",
        "priority": "低", "status": "未着手", "notes": "Phase 2: 2027-01-11",
        "filename": "lambda-dotnet-8",
    },
    {
        "no": 13, "name": "AWS SDK for .NET v3.x", "target": "SDK",
        "current": ".NET SDK v3.x", "eol_date": "2026-03-01",
        "ext_support_end": "2026-06-01", "migration_to": "AWS SDK for .NET v4.x",
        "priority": "高", "status": "未着手", "notes": "メンテナンスモード開始済み、2026-06-01完全EOS",
        "filename": "aws-sdk-dotnet-v3",
        "summary": [
            ("サービス名", "AWS SDK for .NET v3.x"),
            ("EOL対象", "SDK v3.x"),
            ("メンテナンスモード開始", "2026-03-01（セキュリティ修正のみ）"),
            ("完全EOS日", "2026-06-01（全サポート終了）"),
            ("延長サポート", "なし"),
            ("影響範囲", "AWS SDK for .NET v3.xを使用する全アプリケーション"),
            ("セキュリティへの影響", "2026-06-01以降はセキュリティ修正も提供されない"),
            ("代替サービス/移行先", "AWS SDK for .NET v4.x"),
            ("移行の複雑度", "中（API変更・名前空間変更の確認が必要）"),
            ("推定ダウンタイム", "なし（コード変更・再デプロイ）"),
            ("追加コスト", "なし"),
        ],
        "steps": [
            (1, "準備", "使用箇所の洗い出し", "v3.x SDKを参照している全プロジェクトを一覧化", "ソースコードアクセス", "NuGetパッケージ参照で検索", "0.5日"),
            (2, "準備", "v4.xとの差分確認", "Breaking Changesと移行ガイドを確認", "なし", "名前空間・API変更を確認", "1日"),
            (3, "検証", "コード修正・テスト", "v4.x SDKに差し替えてビルド・テスト", "差分確認完了", "CI/CDパイプラインの更新も必要", "2-3日"),
            (4, "実施", "本番デプロイ", "v4.x SDK適用済みアプリケーションをデプロイ", "テスト完了", "段階的なロールアウト推奨", "1日"),
            (5, "確認", "動作確認", "全AWS API呼び出しが正常に動作するか確認", "デプロイ完了", "ログ・メトリクスの監視強化", "0.5日"),
        ],
        "dates": [
            ("メンテナンスモード開始", "2026-03-01", "セキュリティ修正のみ提供"),
            ("完全EOS", "2026-06-01", "全サポート終了"),
            ("推奨対応完了日", "2026-05-01", "完全EOS 1ヶ月前"),
        ],
        "refs": [
            {"title": "AWS SDK for .NET v3 EOS発表", "url": "https://aws.amazon.com/blogs/devops/announcing-the-end-of-support-for-the-aws-sdk-for-net-v3/", "type": "AWSブログ", "desc": "公式EOSアナウンスと移行ガイド"},
        ],
    },
    {
        "no": 14, "name": "ElastiCache Redis 6.x", "target": "エンジンバージョン",
        "current": "Redis 6.x", "eol_date": "2027-01-31",
        "ext_support_end": "2030-01-31", "migration_to": "ElastiCache for Valkey 8.0 / Redis 7.x",
        "priority": "低", "status": "未着手", "notes": "Extended Support: +80%プレミアム",
        "filename": "elasticache-redis-6",
    },
    {
        "no": 15, "name": "RDS/Aurora PostgreSQL 14.x", "target": "メジャーバージョン",
        "current": "PostgreSQL 14.x", "eol_date": "2027-02-28",
        "ext_support_end": "2030-02-28", "migration_to": "PostgreSQL 16/17以上",
        "priority": "低", "status": "未着手", "notes": "Extended Support: $0.100/vCPU-hr (Y1-2)",
        "filename": "rds-aurora-postgresql-14",
    },
    {
        "no": 16, "name": "RDS MySQL 5.7 (延長サポート)", "target": "延長サポート",
        "current": "MySQL 5.7", "eol_date": "2027-02-28",
        "ext_support_end": "2027-02-28", "migration_to": "MySQL 8.0 → 8.4 LTS（段階移行）",
        "priority": "高", "status": "未着手", "notes": "Year 3料金適用中、強制アップグレード間近",
        "filename": "rds-mysql-5.7-extended",
        "summary": [
            ("サービス名", "Amazon RDS for MySQL 5.7（Extended Support）"),
            ("EOL対象バージョン", "MySQL 5.7"),
            ("標準サポート終了日", "2024-02-29（終了済み）"),
            ("延長サポート終了日", "2027-02-28"),
            ("現在の料金ティア", "Year 3: $0.200/vCPU-hr（2026-03-01〜）"),
            ("影響範囲", "延長サポート終了後は強制メジャーバージョンアップグレード"),
            ("セキュリティへの影響", "Extended Supportで重要なCVEパッチのみ提供中"),
            ("代替サービス/移行先", "MySQL 8.0 → MySQL 8.4 LTS（段階移行が必要）"),
            ("移行の複雑度", "高（5.7→8.0で大幅な変更あり、さらに8.0→8.4の2段階）"),
            ("推定ダウンタイム", "各ステップでBlue/Greenデプロイ使用可、合計数時間"),
            ("追加コスト", "現在Year 3料金適用中: $0.200/vCPU-hr"),
        ],
        "steps": [
            (1, "準備", "対象インスタンスの棚卸し", "MySQL 5.7を使用中の全RDSインスタンスを一覧化", "AWSコンソールアクセス", "延長サポート費用も確認", "0.5日"),
            (2, "準備", "5.7→8.0互換性の確認", "MySQL 5.7→8.0間の非互換項目を調査（予約語、データ型、認証）", "なし", "直接5.7→8.4は不可、8.0経由が必須", "2日"),
            (3, "検証", "検証環境で5.7→8.0アップグレード", "Blue/Greenデプロイで5.7→8.0を実施・テスト", "検証環境準備", "アプリケーション互換性テストを徹底", "3日"),
            (4, "実施", "本番5.7→8.0アップグレード", "本番でBlue/Greenデプロイを実施", "テスト完了", "ロールバック手順を確認", "1-2日"),
            (5, "検証", "検証環境で8.0→8.4アップグレード", "Blue/Greenデプロイで8.0→8.4を実施・テスト", "8.0アップグレード完了", "mysql_native_password廃止に注意", "2日"),
            (6, "実施", "本番8.0→8.4アップグレード", "本番でBlue/Greenデプロイを実施", "テスト完了", "認証プラグイン変更の確認", "1-2日"),
            (7, "確認", "最終動作確認", "全アプリケーションの接続・クエリ動作を確認", "全アップグレード完了", "パフォーマンス監視", "1日"),
        ],
        "dates": [
            ("標準サポート終了日", "2024-02-29", "終了済み"),
            ("延長サポート Year 1-2", "2024-03-01", "$0.100/vCPU-hr（終了済み）"),
            ("延長サポート Year 3（現在）", "2026-03-01", "$0.200/vCPU-hr（現在適用中）"),
            ("延長サポート終了日", "2027-02-28", "強制メジャーバージョンアップグレード"),
            ("推奨対応開始日", "2026-03-09", "今すぐ開始を推奨"),
            ("推奨対応完了日", "2027-01-28", "EOL 1ヶ月前"),
        ],
        "refs": [
            {"title": "RDS MySQL 5.7 EOL発表", "url": "https://repost.aws/articles/ARWm1Gv0vJTIKCblhWhPXjWg/announcement-amazon-rds-for-mysql-5-7-will-reach-end-of-standard-support-on-february-29-2024", "type": "AWS re:Post", "desc": "公式EOLアナウンス"},
            {"title": "MySQL 5.7 Extended Support自動登録", "url": "https://aws.amazon.com/blogs/aws/your-mysql-5-7-and-postgresql-11-databases-will-be-automatically-enrolled-into-amazon-rds-extended-support/", "type": "AWSブログ", "desc": "自動登録の詳細"},
            {"title": "RDS Extended Support導入", "url": "https://aws.amazon.com/blogs/database/introducing-amazon-rds-extended-support-for-mysql-databases-on-amazon-aurora-and-amazon-rds/", "type": "AWSブログ", "desc": "Extended Support機能の概要"},
            {"title": "RDS Extended Support", "url": "https://docs.aws.amazon.com/AmazonRDS/latest/UserGuide/extended-support.html", "type": "AWS公式ドキュメント", "desc": "Extended Supportの全般ドキュメント"},
        ],
    },
    {
        "no": 17, "name": "Aurora MySQL 2.x (5.7互換) 延長サポート", "target": "延長サポート",
        "current": "Aurora MySQL 2.x", "eol_date": "2027-02-28",
        "ext_support_end": "2027-02-28", "migration_to": "Aurora MySQL 3.x (8.0互換)",
        "priority": "高", "status": "未着手", "notes": "標準サポート終了済み、延長サポート中",
        "filename": "aurora-mysql-2.x-extended",
    },
    {
        "no": 18, "name": "Lambda 2027年EOLランタイム群", "target": "ランタイム",
        "current": "Node.js 22, Python 3.11, Java 8/11/17, Ruby 3.3",
        "eol_date": "2027-03〜2027-06",
        "ext_support_end": "N/A", "migration_to": "各言語の次期バージョン",
        "priority": "低", "status": "未着手", "notes": "2027年前半〜中盤に順次EOL",
        "filename": "lambda-2027-runtimes",
        "summary": [
            ("サービス名", "AWS Lambda 2027年EOLランタイム群"),
            ("対象ランタイム", "Node.js 22 / Python 3.11 / Java 8(AL2), 11, 17 / Ruby 3.3"),
            ("Node.js 22 Phase 1", "2027-04-30"),
            ("Ruby 3.3 Phase 1", "2027-03-31"),
            ("Python 3.11 Phase 1", "2027-06-30"),
            ("Java 8(AL2)/11/17 Phase 1", "2027-06-30"),
            ("延長サポート", "なし（Lambdaランタイムに延長サポートはない）"),
            ("影響範囲", "対象ランタイムを使用する全Lambda関数"),
            ("セキュリティへの影響", "セキュリティパッチの提供停止"),
            ("移行先 (Node.js 22)", "Node.js次期LTS"),
            ("移行先 (Python 3.11)", "Python 3.12 / 3.13"),
            ("移行先 (Java 8/11/17)", "Java 21（AL2023ベース）"),
            ("移行先 (Ruby 3.3)", "Ruby 3.4以降"),
            ("移行の複雑度", "低〜中（依存ライブラリの互換性確認が必要）"),
        ],
        "steps": [
            (1, "準備", "対象Lambda関数の棚卸し", "2027年EOL対象ランタイムを使用中の全Lambda関数を一覧化", "AWSコンソールアクセス", "全リージョン・全アカウントを確認", "1日"),
            (2, "準備", "移行先バージョンの選定", "各言語の次期バージョンの安定性・Lambda対応状況を確認", "対象一覧", "AWS公式のランタイムサポート表を確認", "0.5日"),
            (3, "準備", "依存ライブラリの互換性確認", "各関数の依存ライブラリが移行先で動作するか確認", "移行先決定", "ネイティブライブラリ・レイヤーに注意", "2日"),
            (4, "検証", "検証環境でのテスト", "移行先ランタイムで関数をデプロイしテスト実行", "互換性確認完了", "Java: AL2→AL2023の変更にも注意", "2-3日"),
            (5, "実施", "ランタイム更新・デプロイ", "Lambda関数のランタイムを更新してデプロイ", "テスト完了", "IaCテンプレートも更新", "1日/言語"),
            (6, "確認", "動作確認", "本番環境での動作確認・ログ監視", "デプロイ完了", "段階的にロールアウト推奨", "1日"),
        ],
        "dates": [
            ("Ruby 3.3 Phase 1", "2027-03-31", "新規作成不可"),
            ("Node.js 22 Phase 1", "2027-04-30", "新規作成不可"),
            ("Python 3.11 Phase 1", "2027-06-30", "新規作成不可"),
            ("Java 8(AL2)/11/17 Phase 1", "2027-06-30", "新規作成不可"),
            ("Ruby 3.3 Phase 2", "2027-05-31", "更新不可"),
            ("Node.js 22 Phase 2", "2027-07-01", "更新不可"),
            ("Python 3.11 Phase 2", "2027-08-31", "更新不可"),
            ("Java 8(AL2)/11/17 Phase 2", "2027-08-31", "更新不可"),
        ],
        "refs": [
            {"title": "Lambda ランタイム一覧", "url": "https://docs.aws.amazon.com/lambda/latest/dg/lambda-runtimes.html", "type": "AWS公式ドキュメント", "desc": "全ランタイムのサポートスケジュール"},
            {"title": "AWS Lambda EOL一覧", "url": "https://endoflife.date/aws-lambda", "type": "サードパーティ", "desc": "全ランタイムのEOLスケジュール"},
            {"title": "Lambdaランタイムアップグレード管理", "url": "https://aws.amazon.com/blogs/compute/managing-aws-lambda-runtime-upgrades/", "type": "AWSブログ", "desc": "ランタイムアップグレードのベストプラクティス"},
        ],
    },
    {
        "no": 19, "name": "Windows Server 2012/2012 R2 (ESU)", "target": "OS（ESU期間）",
        "current": "Windows Server 2012 / 2012 R2", "eol_date": "2026-10-13",
        "ext_support_end": "2026-10-13", "migration_to": "Windows Server 2022 / 2025",
        "priority": "高", "status": "未着手", "notes": "ESU Year 3終了、セキュリティパッチ完全停止",
        "filename": "windows-server-2012",
        "summary": [
            ("サービス名", "Windows Server 2012 / 2012 R2（EC2上）"),
            ("EOL対象", "Windows Server 2012 および 2012 R2"),
            ("メインストリームサポート終了", "2018-10-09"),
            ("延長サポート終了", "2023-10-10（終了済み）"),
            ("ESU Year 1終了", "2024-10-08"),
            ("ESU Year 2終了", "2025-10-14"),
            ("ESU Year 3終了（最終）", "2026-10-13"),
            ("影響範囲", "EC2上のWindows Server 2012/2012 R2インスタンス全て\nセキュリティパッチの完全停止"),
            ("セキュリティへの影響", "ESU終了後はCritical/Importantのセキュリティ更新も提供されない"),
            ("代替サービス/移行先", "Windows Server 2022 または Windows Server 2025"),
            ("移行の複雑度", "高（OSレベルの変更、アプリケーション互換性テスト必要）"),
            ("推定ダウンタイム", "新規インスタンス作成→切り替え方式で計画的に対応"),
            ("追加コスト", "ESU: 有償（Azure移行で無償提供あり）\nAWS上では有償購入が必要"),
        ],
        "steps": [
            (1, "準備", "対象インスタンスの棚卸し", "Windows Server 2012/2012 R2のEC2インスタンスを一覧化", "AWSコンソールアクセス", "SSM Inventoryも活用", "1日"),
            (2, "準備", "アプリケーション互換性調査", "稼働中アプリケーションのWindows Server 2022/2025対応状況を確認", "対象一覧", ".NET Framework/IISバージョンの互換性に注意", "3日"),
            (3, "準備", "移行先AMIの選定", "Windows Server 2022/2025のAMI IDを確認", "なし", "ライセンス形態（License Included）を確認", "0.5日"),
            (4, "検証", "検証環境の構築・テスト", "新OSのEC2インスタンスにアプリケーションをデプロイしテスト", "AMI選定完了", "ドメイン参加・セキュリティポリシーの再設定", "5日"),
            (5, "実施", "本番環境の移行", "新OSインスタンスを構築し、DNSやロードバランサーで切り替え", "テスト完了", "データ移行・バックアップ確認", "2-3日"),
            (6, "確認", "動作確認・監視", "移行後の動作確認とモニタリング", "切り替え完了", "旧インスタンスは一定期間保持", "1日"),
            (7, "確認", "旧インスタンスの廃止", "問題なければ旧Windows Server 2012インスタンスを停止・削除", "動作確認完了", "AMIバックアップを取得してから削除", "0.5日"),
        ],
        "dates": [
            ("メインストリームサポート終了", "2018-10-09", "終了済み"),
            ("延長サポート終了", "2023-10-10", "終了済み"),
            ("ESU Year 1終了", "2024-10-08", "終了済み"),
            ("ESU Year 2終了", "2025-10-14", "終了済み"),
            ("ESU Year 3終了（最終）", "2026-10-13", "この日以降セキュリティパッチ完全停止"),
            ("推奨対応開始日", "2026-07-13", "ESU終了3ヶ月前"),
            ("推奨対応完了日", "2026-09-13", "ESU終了1ヶ月前"),
        ],
        "refs": [
            {"title": "Windows Server 2012 ESU概要", "url": "https://learn.microsoft.com/en-us/windows-server/get-started/extended-security-updates-overview", "type": "Microsoft公式", "desc": "ESUの概要と購入方法"},
            {"title": "Microsoft ESU FAQ", "url": "https://learn.microsoft.com/en-us/lifecycle/faq/extended-security-updates", "type": "Microsoft公式", "desc": "ESUに関するFAQ"},
            {"title": "ESU購入ページ", "url": "https://www.microsoft.com/en-us/windows-server/extended-security-updates", "type": "Microsoft公式", "desc": "ESUの購入と料金"},
            {"title": "AWS Windows Server 2012 EOS対応", "url": "https://aws.amazon.com/blogs/modernizing-with-aws/know-your-aws-options-for-microsoft-windows-server-2012-end-of-support/", "type": "AWSブログ", "desc": "AWSでのWindows Server 2012 EOSオプション"},
            {"title": "AWS Trusted Advisor Windows EOSチェック", "url": "https://aws.amazon.com/about-aws/whats-new/2022/09/aws-trusted-advisor-check-microsoft-windows-server-end-of-support-amazon-ec2/", "type": "AWS What's New", "desc": "Trusted AdvisorでのEOSチェック機能"},
            {"title": "Windows Server EOL一覧", "url": "https://endoflife.date/windows-server", "type": "サードパーティ", "desc": "全バージョンのEOLスケジュール"},
        ],
    },
    {
        "no": 20, "name": "Windows Server 2016", "target": "OS（延長サポート）",
        "current": "Windows Server 2016", "eol_date": "2027-01-12",
        "ext_support_end": "2027-01-12", "migration_to": "Windows Server 2022 / 2025",
        "priority": "中", "status": "未着手", "notes": "延長サポート終了、ESU購入で最大2030年まで延長可能",
        "filename": "windows-server-2016",
        "summary": [
            ("サービス名", "Windows Server 2016（EC2上）"),
            ("EOL対象", "Windows Server 2016"),
            ("メインストリームサポート終了", "2022-01-11（終了済み）"),
            ("延長サポート終了", "2027-01-12"),
            ("ESU（有償）", "あり（2027-01-13〜最大2030-01-12、年単位購入）"),
            ("影響範囲", "EC2上のWindows Server 2016インスタンス全て"),
            ("セキュリティへの影響", "延長サポート終了後はセキュリティ更新なし（ESU購入除く）"),
            ("代替サービス/移行先", "Windows Server 2022 または Windows Server 2025"),
            ("移行の複雑度", "中〜高（OS世代の変更、機能差異の確認必要）"),
            ("推定ダウンタイム", "新規インスタンス作成→切り替え方式で計画的に対応"),
            ("追加コスト", "移行しない場合: ESU購入費用\n移行する場合: 新規インスタンス構築・テスト工数"),
        ],
        "steps": [
            (1, "準備", "対象インスタンスの棚卸し", "Windows Server 2016のEC2インスタンスを一覧化", "AWSコンソールアクセス", "SSM Inventoryも活用", "1日"),
            (2, "準備", "アプリケーション互換性調査", "稼働中アプリケーションのWindows Server 2022/2025対応状況を確認", "対象一覧", ".NET Framework/IIS/SQLバージョンの互換性", "2日"),
            (3, "準備", "移行先AMIの選定", "Windows Server 2022/2025のAMI IDを確認", "なし", "ライセンス形態確認", "0.5日"),
            (4, "検証", "検証環境の構築・テスト", "新OSインスタンスにアプリケーションをデプロイしテスト", "AMI選定完了", "GPO・セキュリティポリシーの再適用", "3-5日"),
            (5, "実施", "本番環境の移行", "新OSインスタンスを構築し段階的に切り替え", "テスト完了", "DNS/ELBでの切り替え", "2日"),
            (6, "確認", "動作確認・旧インスタンス廃止", "移行後確認、旧インスタンスのAMIバックアップ後に廃止", "切り替え完了", "一定期間並行稼働推奨", "1日"),
        ],
        "dates": [
            ("メインストリームサポート終了", "2022-01-11", "終了済み"),
            ("延長サポート終了", "2027-01-12", "セキュリティ更新停止"),
            ("ESU Year 1（有償）", "2027-01-13", "Critical/Importantのセキュリティ更新のみ"),
            ("ESU Year 2", "2028-01-12", ""),
            ("ESU Year 3（最終）", "2030-01-12", "ESU終了"),
            ("推奨対応開始日", "2026-10-12", "延長サポート終了3ヶ月前"),
            ("推奨対応完了日", "2026-12-12", "延長サポート終了1ヶ月前"),
        ],
        "refs": [
            {"title": "Windows Server 2016 ライフサイクル", "url": "https://learn.microsoft.com/en-us/lifecycle/products/windows-server-2016", "type": "Microsoft公式", "desc": "公式ライフサイクル情報"},
            {"title": "Windows Server 2016 EOL計画ガイド", "url": "https://techcommunity.microsoft.com/blog/windows-itpro-blog/plan-for-windows-server-2016-and-windows-10-2016-ltsb-end-of-support/4496136", "type": "Microsoft Tech Community", "desc": "EOL対応計画の公式ガイド"},
            {"title": "Microsoft ESU FAQ", "url": "https://learn.microsoft.com/en-us/lifecycle/faq/extended-security-updates", "type": "Microsoft公式", "desc": "ESUに関するFAQ"},
            {"title": "AWS Windows Server EOS対応", "url": "https://aws.amazon.com/blogs/modernizing-with-aws/its-end-of-support-time-again-are-your-windows-servers-secure/", "type": "AWSブログ", "desc": "AWSでのWindows Server EOSオプション"},
            {"title": "Windows Server EOL一覧", "url": "https://endoflife.date/windows-server", "type": "サードパーティ", "desc": "全バージョンのEOLスケジュール"},
        ],
    },
]

# For services without detailed data, generate minimal reports
MINIMAL_SERVICES = {
    5: {  # Aurora PostgreSQL 13.x
        "summary": [
            ("サービス名", "Amazon Aurora PostgreSQL 13.x"),
            ("EOL対象バージョン", "PostgreSQL 13.x"),
            ("コミュニティEOL", "2025-11"),
            ("AWS標準サポート終了日", "2026-02-28"),
            ("延長サポート", "あり（2026-03-01〜2029-02-28、最大3年）"),
            ("延長サポート費用", "Year 1-2: $0.100/vCPU-hr\nYear 3: $0.200/vCPU-hr"),
            ("影響範囲", "標準サポート終了後は新しいマイナーバージョンリリースなし"),
            ("セキュリティへの影響", "Extended Supportで重要なCVEパッチのみ提供"),
            ("代替サービス/移行先", "Aurora PostgreSQL 16以上"),
            ("移行の複雑度", "中（Blue/Greenデプロイで最小ダウンタイム対応可能）"),
            ("推定ダウンタイム", "Blue/Greenデプロイ: 数分"),
            ("追加コスト", "Extended Support: $0.100〜0.200/vCPU-hr"),
        ],
        "steps": [
            (1, "準備", "対象クラスターの棚卸し", "Aurora PG 13.xを使用中の全クラスターを一覧化", "AWSコンソールアクセス", "Reader/Writer両方確認", "0.5日"),
            (2, "準備", "互換性の事前確認", "PG 13→16間の非互換項目を調査", "対象一覧", "拡張機能の互換性確認", "1日"),
            (3, "検証", "Blue/Greenデプロイテスト", "検証環境でアップグレードテスト", "検証環境準備", "クエリ動作確認", "2日"),
            (4, "実施", "本番Blue/Greenデプロイ", "本番クラスターのアップグレード", "テスト完了", "スイッチオーバータイミング計画", "1日"),
            (5, "確認", "動作確認", "パフォーマンス検証", "スイッチオーバー完了", "ロールバック用に旧環境保持", "1日"),
        ],
        "dates": [
            ("コミュニティEOL", "2025-11-01", "PostgreSQLコミュニティサポート終了"),
            ("AWS標準サポート終了日", "2026-02-28", "Extended Support自動登録"),
            ("延長サポート終了日", "2029-02-28", "強制アップグレード"),
            ("推奨対応開始日", "2025-11-28", "EOL 3ヶ月前"),
            ("推奨対応完了日", "2026-01-28", "EOL 1ヶ月前"),
        ],
        "refs": [
            {"title": "Aurora PostgreSQL 13 EOL発表", "url": "https://repost.aws/articles/ARxQSsnCAlS6OhFm7M10vwjA/announcement-amazon-aurora-postgresql-13-x-end-of-standard-support-is-february-28-2026", "type": "AWS re:Post", "desc": "公式EOLアナウンス"},
            {"title": "Aurora PostgreSQLリリースカレンダー", "url": "https://docs.aws.amazon.com/AmazonRDS/latest/AuroraPostgreSQLReleaseNotes/aurorapostgresql-release-calendar.html", "type": "AWS公式ドキュメント", "desc": "サポートスケジュール"},
            {"title": "Aurora Extended Support", "url": "https://docs.aws.amazon.com/AmazonRDS/latest/AuroraUserGuide/extended-support.html", "type": "AWS公式ドキュメント", "desc": "Extended Supportの概要"},
            {"title": "PG 13アップグレード戦略", "url": "https://aws.amazon.com/blogs/database/strategies-for-upgrading-amazon-aurora-postgresql-and-amazon-rds-for-postgresql-from-version-13/", "type": "AWSブログ", "desc": "アップグレード方法"},
            {"title": "Aurora Blue/Greenデプロイ", "url": "https://aws.amazon.com/blogs/database/implement-a-rollback-strategy-for-amazon-aurora-postgresql-upgrades-using-amazon-rds-blue-green-deployments/", "type": "AWSブログ", "desc": "ロールバック戦略"},
        ],
    },
    6: {  # Lambda Ruby 3.2
        "summary": [
            ("サービス名", "AWS Lambda - Ruby 3.2 ランタイム"),
            ("EOL対象", "Ruby 3.2 ランタイム"),
            ("Phase 1（新規作成不可）", "2026-03-31"),
            ("Phase 2（更新不可）", "2026-09-30"),
            ("延長サポート", "なし"),
            ("影響範囲", "Ruby 3.2を使用する全Lambda関数"),
            ("代替サービス/移行先", "Ruby 3.3 / 3.4"),
            ("移行の複雑度", "低"),
        ],
        "steps": [
            (1, "準備", "対象関数の洗い出し", "Ruby 3.2を使用中の全Lambda関数を一覧化", "AWSコンソールアクセス", "全リージョン確認", "0.5日"),
            (2, "検証", "互換性テスト", "Ruby 3.3/3.4でのテスト実行", "対象一覧", "Gem互換性確認", "1日"),
            (3, "実施", "ランタイム更新", "Ruby 3.3/3.4に変更してデプロイ", "テスト完了", "IaCテンプレート更新", "0.5日/関数"),
        ],
        "dates": [
            ("Phase 1（新規作成不可）", "2026-03-31", ""),
            ("Phase 2（更新不可）", "2026-09-30", ""),
        ],
        "refs": [
            {"title": "Lambda ランタイム一覧", "url": "https://docs.aws.amazon.com/lambda/latest/dg/lambda-runtimes.html", "type": "AWS公式ドキュメント", "desc": "サポートスケジュール"},
        ],
    },
    7: {  # Lambda Node.js 20.x
        "summary": [
            ("サービス名", "AWS Lambda - Node.js 20.x ランタイム"),
            ("EOL対象", "Node.js 20.x ランタイム"),
            ("Phase 1（セキュリティパッチ停止）", "2026-04-30"),
            ("Phase 2a（新規作成不可）", "2026-06-01"),
            ("Phase 2b（更新不可）", "2026-07-01"),
            ("延長サポート", "なし"),
            ("影響範囲", "Node.js 20.xを使用する全Lambda関数"),
            ("代替サービス/移行先", "Node.js 22"),
            ("移行の複雑度", "低〜中"),
        ],
        "steps": [
            (1, "準備", "対象関数の洗い出し", "Node.js 20.xを使用中の全Lambda関数を一覧化", "AWSコンソールアクセス", "全リージョン確認", "0.5日"),
            (2, "検証", "互換性テスト", "Node.js 22でのテスト実行", "対象一覧", "npm依存パッケージの互換性確認", "1-2日"),
            (3, "実施", "ランタイム更新", "Node.js 22に変更してデプロイ", "テスト完了", "IaCテンプレート更新", "0.5日/関数"),
        ],
        "dates": [
            ("Phase 1（パッチ停止）", "2026-04-30", ""),
            ("Phase 2a（新規作成不可）", "2026-06-01", ""),
            ("Phase 2b（更新不可）", "2026-07-01", ""),
        ],
        "refs": [
            {"title": "Lambda ランタイム一覧", "url": "https://docs.aws.amazon.com/lambda/latest/dg/lambda-runtimes.html", "type": "AWS公式ドキュメント", "desc": "サポートスケジュール"},
            {"title": "Lambda Node.js 20 EOL移行ガイド", "url": "https://www.cloudquery.io/blog/aws-lambda-nodejs-20-eol", "type": "サードパーティ", "desc": "Node.js 22への移行方法"},
        ],
    },
    9: {  # Lambda Custom Runtime AL2
        "summary": [
            ("サービス名", "AWS Lambda - Custom Runtime (provided.al2)"),
            ("EOL対象", "provided.al2 ランタイム"),
            ("Phase 1（パッチ停止）", "2026-07-31"),
            ("Phase 2（更新不可）", "2026-09-30"),
            ("延長サポート", "なし"),
            ("影響範囲", "AL2ベースのカスタムランタイムを使用する全Lambda関数"),
            ("代替サービス/移行先", "provided.al2023"),
            ("移行の複雑度", "中（AL2→AL2023のOS差分に注意）"),
        ],
        "steps": [
            (1, "準備", "対象関数の洗い出し", "provided.al2を使用中の全Lambda関数を一覧化", "AWSコンソールアクセス", "全リージョン確認", "0.5日"),
            (2, "準備", "AL2023との差分確認", "カスタムランタイムのバイナリ・ライブラリの互換性確認", "対象一覧", "glibc等のバージョン差異に注意", "1日"),
            (3, "検証", "AL2023でのビルド・テスト", "provided.al2023用にリビルドしてテスト", "差分確認完了", "ネイティブライブラリの再コンパイルが必要な場合あり", "2日"),
            (4, "実施", "ランタイム変更・デプロイ", "provided.al2023に変更してデプロイ", "テスト完了", "IaCテンプレート更新", "1日"),
        ],
        "dates": [
            ("Phase 1（パッチ停止）", "2026-07-31", "Amazon Linux 2 EOLに連動"),
            ("Phase 2（更新不可）", "2026-09-30", ""),
        ],
        "refs": [
            {"title": "Lambda ランタイム一覧", "url": "https://docs.aws.amazon.com/lambda/latest/dg/lambda-runtimes.html", "type": "AWS公式ドキュメント", "desc": "サポートスケジュール"},
        ],
    },
    11: {  # Lambda Python 3.10
        "summary": [
            ("サービス名", "AWS Lambda - Python 3.10 ランタイム"),
            ("EOL対象", "Python 3.10 ランタイム"),
            ("Phase 1（パッチ停止）", "2026-10-31"),
            ("Phase 2（更新不可）", "2027-01-15"),
            ("延長サポート", "なし"),
            ("影響範囲", "Python 3.10を使用する全Lambda関数"),
            ("代替サービス/移行先", "Python 3.12 / 3.13"),
            ("移行の複雑度", "低〜中"),
        ],
        "steps": [
            (1, "準備", "対象関数の洗い出し", "Python 3.10を使用中の全Lambda関数を一覧化", "AWSコンソールアクセス", "全リージョン確認", "0.5日"),
            (2, "検証", "互換性テスト", "Python 3.12/3.13でのテスト実行", "対象一覧", "依存ライブラリの互換性確認", "1日"),
            (3, "実施", "ランタイム更新", "Python 3.12/3.13に変更してデプロイ", "テスト完了", "IaCテンプレート更新", "0.5日/関数"),
        ],
        "dates": [
            ("Phase 1（パッチ停止）", "2026-10-31", ""),
            ("Phase 2（更新不可）", "2027-01-15", ""),
        ],
        "refs": [
            {"title": "Lambda ランタイム一覧", "url": "https://docs.aws.amazon.com/lambda/latest/dg/lambda-runtimes.html", "type": "AWS公式ドキュメント", "desc": "サポートスケジュール"},
        ],
    },
    12: {  # Lambda .NET 8
        "summary": [
            ("サービス名", "AWS Lambda - .NET 8 ランタイム"),
            ("EOL対象", ".NET 8 ランタイム"),
            ("Phase 1（パッチ停止）", "2026-11-10"),
            ("Phase 2（更新不可）", "2027-01-11"),
            ("延長サポート", "なし"),
            ("影響範囲", ".NET 8を使用する全Lambda関数"),
            ("代替サービス/移行先", ".NET 10"),
            ("移行の複雑度", "中（.NETメジャーバージョン間の変更確認要）"),
        ],
        "steps": [
            (1, "準備", "対象関数の洗い出し", ".NET 8を使用中の全Lambda関数を一覧化", "AWSコンソールアクセス", "全リージョン確認", "0.5日"),
            (2, "検証", "互換性テスト", ".NET 10でのビルド・テスト実行", "対象一覧", "NuGetパッケージ互換性確認", "2日"),
            (3, "実施", "ランタイム更新", ".NET 10に変更してデプロイ", "テスト完了", "IaCテンプレート更新", "1日"),
        ],
        "dates": [
            ("Phase 1（パッチ停止）", "2026-11-10", ""),
            ("Phase 2（更新不可）", "2027-01-11", ""),
        ],
        "refs": [
            {"title": "Lambda ランタイム一覧", "url": "https://docs.aws.amazon.com/lambda/latest/dg/lambda-runtimes.html", "type": "AWS公式ドキュメント", "desc": "サポートスケジュール"},
        ],
    },
    14: {  # ElastiCache Redis 6.x
        "summary": [
            ("サービス名", "Amazon ElastiCache for Redis OSS 6.x"),
            ("EOL対象バージョン", "Redis 6.x"),
            ("標準サポート終了日", "2027-01-31"),
            ("延長サポート", "あり（2027-02-01〜2030-01-31、最大3年）"),
            ("延長サポート費用", "Year 1-2: On-Demand価格の+80%\nYear 3: On-Demand価格の+160%"),
            ("影響範囲", "Redis 6.xを使用する全ElastiCacheクラスター"),
            ("代替サービス/移行先", "ElastiCache for Valkey 8.0 / Redis 7.x"),
            ("移行の複雑度", "低〜中"),
        ],
        "steps": [
            (1, "準備", "対象クラスターの棚卸し", "Redis 6.xを使用中の全クラスターを一覧化", "AWSコンソールアクセス", "全リージョン確認", "0.5日"),
            (2, "検証", "アップグレードテスト", "検証環境でValkey/Redis 7.xへのアップグレードテスト", "検証環境準備", "API互換性確認", "1日"),
            (3, "実施", "インプレースアップグレード", "本番クラスターをValkey/Redis 7.xにアップグレード", "テスト完了", "ダウンタイム最小化のためレプリカ追加推奨", "1日"),
        ],
        "dates": [
            ("標準サポート終了日", "2027-01-31", "Extended Support自動登録"),
            ("延長サポート終了日", "2030-01-31", "強制アップグレード"),
            ("推奨対応開始日", "2026-10-31", "EOL 3ヶ月前"),
        ],
        "refs": [
            {"title": "ElastiCache Extended Support対象バージョン", "url": "https://docs.aws.amazon.com/AmazonElastiCache/latest/dg/extended-support-versions.html", "type": "AWS公式ドキュメント", "desc": "対象バージョン一覧"},
            {"title": "ElastiCache Extended Support", "url": "https://docs.aws.amazon.com/AmazonElastiCache/latest/dg/extended-support.html", "type": "AWS公式ドキュメント", "desc": "概要"},
        ],
    },
    15: {  # RDS/Aurora PostgreSQL 14.x
        "summary": [
            ("サービス名", "Amazon RDS / Aurora PostgreSQL 14.x"),
            ("EOL対象バージョン", "PostgreSQL 14.x"),
            ("コミュニティEOL", "2026-11"),
            ("AWS標準サポート終了日", "2027-02-28"),
            ("延長サポート", "あり（2027-03-01〜2030-02-28、最大3年）"),
            ("延長サポート費用", "Year 1-2: $0.100/vCPU-hr\nYear 3: $0.200/vCPU-hr"),
            ("影響範囲", "RDS/Aurora PG 14.xの全インスタンス/クラスター"),
            ("代替サービス/移行先", "PostgreSQL 16 / 17以上"),
            ("移行の複雑度", "中"),
        ],
        "steps": [
            (1, "準備", "対象の棚卸し", "PG 14.xを使用中の全RDS/Auroraリソースを一覧化", "AWSコンソールアクセス", "RDSとAurora両方確認", "0.5日"),
            (2, "検証", "アップグレードテスト", "Blue/GreenデプロイでPG 16/17へのアップグレードテスト", "検証環境準備", "互換性確認", "2日"),
            (3, "実施", "本番アップグレード", "Blue/Greenデプロイで本番アップグレード", "テスト完了", "スイッチオーバー計画", "1日"),
        ],
        "dates": [
            ("コミュニティEOL", "2026-11-01", "PostgreSQLコミュニティサポート終了"),
            ("AWS標準サポート終了日", "2027-02-28", "Extended Support自動登録"),
            ("延長サポート終了日", "2030-02-28", "強制アップグレード"),
            ("推奨対応開始日", "2026-11-28", "EOL 3ヶ月前"),
        ],
        "refs": [
            {"title": "RDS PostgreSQLリリースカレンダー", "url": "https://docs.aws.amazon.com/AmazonRDS/latest/PostgreSQLReleaseNotes/postgresql-release-calendar.html", "type": "AWS公式ドキュメント", "desc": "サポートスケジュール"},
            {"title": "Aurora PostgreSQLリリースカレンダー", "url": "https://docs.aws.amazon.com/AmazonRDS/latest/AuroraPostgreSQLReleaseNotes/aurorapostgresql-release-calendar.html", "type": "AWS公式ドキュメント", "desc": "サポートスケジュール"},
        ],
    },
    17: {  # Aurora MySQL 2.x Extended
        "summary": [
            ("サービス名", "Amazon Aurora MySQL 2.x (MySQL 5.7互換) Extended Support"),
            ("EOL対象バージョン", "Aurora MySQL 2.x"),
            ("標準サポート終了日", "2024-10-31（終了済み）"),
            ("延長サポート終了日", "2027-02-28"),
            ("延長サポート費用", "$0.100/vCPU-hr"),
            ("影響範囲", "延長サポート終了後は強制メジャーバージョンアップグレード"),
            ("代替サービス/移行先", "Aurora MySQL 3.x (MySQL 8.0互換)"),
            ("移行の複雑度", "高（MySQL 5.7→8.0の大幅な変更）"),
        ],
        "steps": [
            (1, "準備", "対象クラスターの棚卸し", "Aurora MySQL 2.xを使用中の全クラスターを一覧化", "AWSコンソールアクセス", "延長サポート費用も確認", "0.5日"),
            (2, "準備", "MySQL 8.0互換性確認", "予約語変更・認証プラグイン・lower_case_table_names設定等を確認", "対象一覧", "v2→v3のBreaking Changes確認", "2日"),
            (3, "準備", "パラメータグループ準備", "MySQL 8.0互換のカスタムパラメータグループを作成", "互換性確認完了", "lower_case_table_namesの一致確認必須", "0.5日"),
            (4, "検証", "Blue/Greenデプロイテスト", "検証環境でv2→v3のアップグレードテスト", "パラメータグループ準備完了", "プリチェック失敗の対処", "3日"),
            (5, "実施", "本番アップグレード", "本番でBlue/Greenデプロイを実施", "テスト完了", "事前スナップショット自動作成される", "1-2日"),
            (6, "確認", "動作確認", "アプリケーション動作・パフォーマンス確認", "アップグレード完了", "ロールバック手順確認", "1日"),
        ],
        "dates": [
            ("標準サポート終了日", "2024-10-31", "終了済み"),
            ("延長サポート課金開始", "2024-12-01", "Extended Support料金発生中"),
            ("延長サポート終了日", "2027-02-28", "強制アップグレード"),
            ("推奨対応開始日", "2026-03-09", "今すぐ開始を推奨"),
        ],
        "refs": [
            {"title": "Aurora MySQL 2 EOL発表", "url": "https://repost.aws/articles/ARzYGTjZP8Tbelta9ZRtYRxw/announcement-amazon-aurora-mysql-2-x-end-of-standard-support-is-on-october-31-2024", "type": "AWS re:Post", "desc": "公式EOLアナウンス"},
            {"title": "Aurora MySQL v2 EOL準備", "url": "https://docs.aws.amazon.com/AmazonRDS/latest/AuroraUserGuide/Aurora.MySQL57.EOL.html", "type": "AWS公式ドキュメント", "desc": "EOL準備ガイド"},
            {"title": "Aurora MySQL v2→v3チェックリスト Part 1", "url": "https://aws.amazon.com/blogs/database/amazon-aurora-mysql-version-2-with-mysql-5-7-compatibility-to-version-3-with-mysql-8-0-compatibility-upgrade-checklist-part-1/", "type": "AWSブログ", "desc": "アップグレードチェックリスト"},
            {"title": "Aurora MySQL v2→v3チェックリスト Part 2", "url": "https://aws.amazon.com/blogs/database/amazon-aurora-mysql-version-2-with-mysql-5-7-compatibility-to-version-3-with-mysql-8-0-compatibility-upgrade-checklist-part-2/", "type": "AWSブログ", "desc": "アップグレードチェックリスト続き"},
        ],
    },
}


def generate_individual_report(svc):
    """Generate individual service EOL report Excel."""
    no = svc["no"]
    wb = Workbook()

    # Get detailed data (from main or minimal)
    detail = MINIMAL_SERVICES.get(no, svc)
    summary_data = detail.get("summary", svc.get("summary", [
        ("サービス名", svc["name"]),
        ("EOL対象", svc["target"]),
        ("EOL日程", svc["eol_date"]),
        ("延長サポート終了日", svc.get("ext_support_end", "N/A")),
        ("移行先", svc["migration_to"]),
        ("備考", svc["notes"]),
    ]))
    steps_data = detail.get("steps", svc.get("steps", [
        (1, "準備", "対象リソースの棚卸し", "対象サービスを使用中のリソースを一覧化する", "AWSコンソールアクセス", "-", "0.5日"),
        (2, "検証", "互換性テスト", "移行先バージョンでの動作テストを実施", "棚卸し完了", "-", "1-2日"),
        (3, "実施", "移行実施", "本番環境で移行を実施", "テスト完了", "ロールバック手順確認", "1日"),
    ]))
    dates_data = detail.get("dates", svc.get("dates", [
        ("EOL日程", svc["eol_date"], ""),
    ]))
    refs_data = detail.get("refs", svc.get("refs", [
        {"title": "AWS Product Lifecycle", "url": "https://aws.amazon.com/products/lifecycle/", "type": "AWS公式", "desc": "AWS製品ライフサイクル一覧"},
    ]))

    # Sheet 1: Summary
    ws1 = wb.active
    write_kv_sheet(ws1, "調査内容まとめ", summary_data)

    # Sheet 2: EOL Steps
    ws2 = wb.create_sheet()
    write_eol_steps_sheet(ws2, "EOL手順", steps_data)

    # Sheet 3: EOL Dates
    ws3 = wb.create_sheet()
    write_eol_dates_sheet(ws3, "EOL期限", dates_data)

    # Sheet 4: References
    ws4 = wb.create_sheet()
    write_references_sheet(ws4, "参考文献", refs_data)

    # Sheet 5: CloudShell Commands
    cs_commands = get_cloudshell_commands_for_service(svc["filename"])
    if cs_commands:
        ws5 = wb.create_sheet()
        write_cloudshell_sheet(ws5, "CloudShellリソース洗い出し", cs_commands)

    filename = f"{svc['filename']}-eol-report-2026-03-09.xlsx"
    save_wb(wb, filename)


def generate_management_report():
    """Generate management overview Excel."""
    wb = Workbook()

    # Sheet 1: Service List
    ws1 = wb.active
    headers = ["No.", "サービス名", "EOL対象", "現行バージョン", "EOL日程", "延長サポート終了日", "移行先", "優先度", "ステータス", "備考"]
    widths = [6, 30, 20, 25, 15, 18, 30, 10, 12, 40]
    write_table_sheet(ws1, "EOL対象サービス一覧", headers, widths, [], center_cols=[1, 5, 6, 8, 9])

    for r, svc in enumerate(SERVICES, 2):
        d = days_until(svc["eol_date"])
        data = [svc["no"], svc["name"], svc["target"], svc["current"], svc["eol_date"],
                svc.get("ext_support_end", "N/A"), svc["migration_to"], svc["priority"], svc["status"], svc["notes"]]
        for c, val in enumerate(data, 1):
            cell = style_data_cell(ws1, r, c, center=(c in [1, 5, 6, 8, 9]))
            cell.value = val
        # Priority coloring
        p_cell = ws1.cell(row=r, column=8)
        if svc["priority"] == "高":
            p_cell.fill = RED_FILL
        elif svc["priority"] == "中":
            p_cell.fill = YELLOW_FILL
        else:
            p_cell.fill = GREEN_FILL
        # Status coloring
        s_cell = ws1.cell(row=r, column=9)
        if svc["status"] == "未着手":
            s_cell.font = RED_FONT
        elif svc["status"] == "調査中":
            s_cell.font = BLUE_FONT
        elif svc["status"] == "対応中":
            s_cell.font = ORANGE_FONT
        elif svc["status"] == "完了":
            s_cell.font = GREEN_FONT

    # Sheet 2: Task List
    ws2 = wb.create_sheet()
    task_headers = ["No.", "対象サービス", "タスク名", "担当者", "開始予定日", "期限", "ステータス", "依存関係", "備考"]
    task_widths = [6, 25, 35, 15, 15, 15, 12, 30, 40]
    write_table_sheet(ws2, "対応タスク一覧", task_headers, task_widths, [], center_cols=[1, 5, 6, 7])

    task_no = 1
    task_templates = [
        "EOL情報の詳細調査・確認",
        "影響範囲の洗い出し",
        "移行計画の策定",
        "検証環境での移行テスト",
        "本番環境での移行実施",
        "移行後の動作確認",
        "旧リソースのクリーンアップ",
    ]
    for svc in SERVICES:
        for tmpl in task_templates:
            data = [task_no, svc["name"], tmpl, "", "", "", "未着手", "", ""]
            r = task_no + 1
            for c, val in enumerate(data, 1):
                cell = style_data_cell(ws2, r, c, center=(c in [1, 5, 6, 7]))
                cell.value = val
            ws2.cell(row=r, column=7).font = RED_FONT
            task_no += 1

    # Sheet 3: Schedule
    ws3 = wb.create_sheet()
    sch_headers = ["サービス名", "EOL期限", "残日数", "調査完了目標", "検証開始目標", "本番対応目標", "バッファ期間", "備考"]
    sch_widths = [30, 15, 10, 15, 15, 15, 12, 40]
    write_table_sheet(ws3, "スケジュール", sch_headers, sch_widths, [], center_cols=[2, 3, 4, 5, 6, 7])

    for r, svc in enumerate(SERVICES, 2):
        d = days_until(svc["eol_date"])
        remaining = d if d is not None else "N/A"
        data = [svc["name"], svc["eol_date"], remaining, "", "", "", "", svc["notes"]]
        for c, val in enumerate(data, 1):
            cell = style_data_cell(ws3, r, c, center=(c in [2, 3, 4, 5, 6, 7]))
            cell.value = val
        # Color remaining days
        if d is not None:
            days_cell = ws3.cell(row=r, column=3)
            if d < 30:
                days_cell.fill = RED_FILL
            elif d < 90:
                days_cell.fill = YELLOW_FILL
            else:
                days_cell.fill = GREEN_FILL

    # Sheet 4: CloudShell Commands (all categories)
    ws4 = wb.create_sheet()
    all_commands = []
    for cat_key in ["rds", "aurora", "elasticache", "opensearch", "lambda", "ec2", "windows_server", "multi_account"]:
        all_commands.extend(CLOUDSHELL_COMMANDS.get(cat_key, []))
    write_cloudshell_sheet(ws4, "CloudShellリソース洗い出し", all_commands)

    save_wb(wb, "aws-eol-management-2026-03-09.xlsx")


# ============================================================
# Main
# ============================================================
if __name__ == "__main__":
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    print("=== AWS EOL Reports Generator ===")
    print(f"Output: {OUTPUT_DIR}\n")

    print("Generating individual service reports...")
    for svc in SERVICES:
        generate_individual_report(svc)

    print("\nGenerating management report...")
    generate_management_report()

    print(f"\nDone! Generated {len(SERVICES) + 1} files.")
